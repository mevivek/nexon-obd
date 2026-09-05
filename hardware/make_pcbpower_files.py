"""Emit a PCB Power style Pick and Place file for one variant.

  python3 make_pcbpower_files.py xiao 2

Their format is not one you would guess, so it is copied from their own
Sample_PickandPlace.xlsx rather than inferred:

  row 1   a title cell, "SAMPLE PICK AND PLACE"
  row 2   headers starting in column B - column A is left EMPTY
  row 3+  data, also from column B
  columns Reference Designator | X | Y | Rotation | Side | Value
  Side    "Top" / "Bottom", capitalised
  X, Y    positive millimetres, board-relative

Written as .xlsx because that is what their sample is, with a .csv beside it in
the same shape in case their importer prefers one.

The lesson this file encodes: read the vendor's sample. Three formats were
guessed at before this one and all three were wrong in some detail.
"""
import csv, io, os, sys

HERE = os.path.dirname(os.path.abspath(__file__))
HALF = 20.0                # 40 x 40 board centred on its own origin
TITLE = 'OBDURATE REV A PICK AND PLACE'
HEADERS = ['Reference Designator', 'X', 'Y', 'Rotation', 'Side', 'Value']


def rows(name):
    src = [l for l in io.open(os.path.join(HERE, name), encoding='utf-8')
           if l.strip() and not l.startswith('#')]
    return list(csv.DictReader(src))


def collect(variant):
    keep = ('common', variant)
    bom = {r['designator']: r for r in rows('bom.csv') if r['variant'] in keep}
    out = []
    for r in rows('placement.csv'):
        if r['variant'] not in keep:
            continue
        ref = r['designator']
        b = bom.get(ref)
        if b is None or b['asm'].strip() != 'y' or b['dnp'].strip() == 'y':
            continue
        out.append([
            ref,
            round(float(r['x_mm']) + HALF, 3),
            round(float(r['y_mm']) + HALF, 3),
            int(float(r['rotation_deg'])),
            'Top' if r['side'] == 'top' else 'Bottom',
            b['value_or_mpn'].strip(),
        ])
    return out


def powerbom(variant, path):
    """PowerBoM wants ONE ROW PER UNIQUE MPN with the designators grouped -
    it rejects the file outright with "Duplicate MPN found" otherwise. That is
    the opposite of the per-designator BOM LionCircuits asked for, which is why
    this is a separate output rather than the same file renamed.

    Field names below are spelled exactly as their mapper lists them, including
    "Manufacture Part Number", so the columns auto-assign."""
    keep = ('common', variant)
    bom = [r for r in rows('bom.csv') if r['variant'] in keep]
    plc = {r['designator'] for r in rows('placement.csv') if r['variant'] in keep}

    groups = {}
    for r in bom:
        ref = r['designator']
        if r['asm'].strip() != 'y' or ref not in plc:
            continue
        if r['dnp'].strip() == 'y':
            continue            # do not source what is not fitted
        key = r['mpn'].strip() or r['value_or_mpn'].strip()
        g = groups.setdefault(key, {'refs': [], 'row': r})
        g['refs'].append(ref)

    def sortkey(ref):
        head = ref.rstrip('0123456789')
        tail = ref[len(head):]
        return (head, int(tail) if tail else 0)

    cols = ['Sr#', 'Ref Designator', 'Quantity/PCB', 'Description', 'Value',
            'Mfg Name', 'Manufacture Part Number', 'Footprint', 'Note']
    ordered = sorted(groups.items(),
                     key=lambda kv: sortkey(sorted(kv[1]['refs'], key=sortkey)[0]))
    with io.open(path, 'w', encoding='utf-8', newline='') as fh:
        w = csv.writer(fh, lineterminator='\n')
        w.writerow(cols)
        for i, (mpn, g) in enumerate(ordered, 1):
            r = g['row']
            refs = sorted(g['refs'], key=sortkey)
            note = r['note']
            keepn = [k for k in ('NO SUBSTITUTION', 'value_tbd', 'REQUIRED',
                                 'consign') if k in note]
            w.writerow([i, ','.join(refs), len(refs),
                        describe_row(r), r['value_or_mpn'].strip(),
                        r['manufacturer'].strip(), mpn,
                        r['package'].strip(), '; '.join(keepn)])
    return len(ordered), sum(len(g['refs']) for _, g in ordered)


TYPES = [('LED', 'LED'), ('CH', 'Choke'), ('TP', 'Test pad'), ('U', 'IC'),
         ('R', 'Resistor'), ('C', 'Capacitor'), ('L', 'Inductor'),
         ('D', 'Diode'), ('Q', 'Transistor'), ('F', 'Fuse'),
         ('J', 'Connector'), ('W', 'Pad')]


def describe_row(r):
    ref = r['designator']
    kind = 'Module' if ref == 'U3' else next(
        (n for p, n in TYPES if ref.startswith(p)), 'Part')
    bits = [kind, r['value_or_mpn'].strip(), r['package'].strip()]
    return ', '.join(b for b in bits if b)


def main(variant, layers):
    data = collect(variant)
    out = os.path.join(HERE, 'gerbers', '%s-%dL-quote' % (variant, layers))
    os.makedirs(out, exist_ok=True)
    stem = os.path.join(out, 'obdurate-revA-%s-%dL-pickandplace' % (variant, layers))

    # .csv in the same shape - note the leading empty column
    with io.open(stem + '.csv', 'w', encoding='utf-8', newline='') as fh:
        w = csv.writer(fh, lineterminator='\n')
        w.writerow([TITLE])
        w.writerow([''] + HEADERS)
        for row in data:
            w.writerow([''] + row)

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Pick and Place'
        ws['A1'] = TITLE
        for i, h in enumerate(HEADERS):
            ws.cell(row=2, column=2 + i, value=h)
        for j, row in enumerate(data):
            for i, v in enumerate(row):
                ws.cell(row=3 + j, column=2 + i, value=v)
        for col, width in zip('ABCDEFG', (4, 22, 10, 10, 10, 10, 26)):
            ws.column_dimensions[col].width = width
        wb.save(stem + '.xlsx')
        made = 'xlsx + csv'
    except ImportError:
        made = 'csv only (openpyxl not installed)'

    sys.stderr.write('%s %dL: %d placements -> %s (%s)\n'
                     % (variant, layers, len(data),
                        os.path.relpath(stem, HERE), made))

    bpath = os.path.join(out, 'obdurate-revA-%s-%dL-powerbom.csv' % (variant, layers))
    uniq, total = powerbom(variant, bpath)
    sys.stderr.write('  PowerBoM: %d unique parts covering %d placements -> %s\n'
                     % (uniq, total, os.path.basename(bpath)))
    sys.stderr.write('  U5 is excluded as DNP. If you want the option of fitting it,'
                     ' buy a few separately - it is not on this quote.\n')
    return 0


if __name__ == '__main__':
    v = sys.argv[1] if len(sys.argv) > 1 else 'xiao'
    n = int(sys.argv[2]) if len(sys.argv) > 2 else 2
    sys.exit(main(v, n))
